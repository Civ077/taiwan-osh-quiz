# -*- coding: utf-8 -*-
"""組裝題庫：讀 batchN_*.py → 驗證 → 洗牌選項 → 產出
   OSH_ENV_QuizBank.xlsx（Laws/Questions/Config/Changelog 四分頁，全部題目）
   questions_all.json（全部）、questions_batchN.json（各批）
   tsv/Questions_batchN.tsv（各批新增列，貼進 Google Sheet 用）
用法：python build_sheet.py            # 全部批次
"""
import re
import json, sys, os, importlib, random, glob, re
from pathlib import Path as _Path
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)

# 法規主檔：law_id → (中文名, 英文名, 版本日期, pcode)
def law_version_from_text(zh):
    """從 法規原文/<名稱>.txt 第一行讀修正日期 → 民國115年6月26日"""
    f = os.path.join(os.path.dirname(HERE), "法規原文", zh + ".txt")
    if not os.path.exists(f): return ""
    m = re.search(r"修正日期：民國\s*(\d+)\s*年\s*(\d+)\s*月\s*(\d+)\s*日", open(f, encoding="utf-8").readline())
    if m: return f"民國{int(m.group(1))}年{int(m.group(2))}月{int(m.group(3))}日"
    m = re.search(r"修正日期：民國\s*(\d+)\s*年\s*(\d+)\s*月", open(f, encoding="utf-8").readline())
    return f"民國{int(m.group(1))}年{int(m.group(2))}月" if m else ""


LAWS = {
 "OSH-01": ("職業安全衛生法","Occupational Safety and Health Act","民國114年12月19日","N0060001"),
 "OSH-02": ("職業安全衛生法施行細則","Enforcement Rules of the OSH Act","民國115年6月26日","N0060002"),
 "OSH-07": ("職業安全衛生設施規則","Occupational Safety and Health Facilities Rules","民國115年6月30日","N0060009"),
 "OSH-13": ("機械設備器具安全標準","Safety Standards for Machinery, Equipment and Tools","民國111年5月11日","N0060034"),
 "OSH-36": ("營造安全衛生設施標準","Construction Safety and Health Facilities Standards","民國115年6月30日","N0060014"),
 "OSH-34": ("缺氧症預防規則","Rules for Prevention of Hypoxia","民國103年6月26日","N0060020"),
 "OSH-15": ("高架作業勞工保護措施標準","Standards for Protective Measures for Laborers in Work at Height","民國103年6月25日","N0060029"),
 "OSH-14": ("高溫作業勞工作息時間標準","Standards for Work and Rest Time of Laborers in High-Temperature Work","民國103年7月1日","N0060007"),
 "OSH-18": ("異常氣壓危害預防標準","Standards for Prevention of Abnormal Pressure Hazards","民國103年6月25日","N0060026"),
 "OSH-12": ("危害性化學品標示及通識規則","Regulations on Labeling and Hazard Communication of Hazardous Chemicals","民國115年8月26日","N0060054"),
 "OSH-24": ("職業安全衛生標示設置準則","Guidelines for Installation of Occupational Safety and Health Signs","民國103年7月2日","N0060023"),
 "OSH-09": ("職業安全衛生管理辦法","Regulations of Occupational Safety and Health Management","民國115年6月29日","N0060027"),
 "OSH-11": ("職業安全衛生教育訓練規則","Occupational Safety and Health Education and Training Rules","民國115年6月25日","N0060010"),
 "OSH-10": ("勞工健康保護規則","Labor Health Protection Rules","民國115年6月26日","N0060022"),
 "OSH-08": ("危險性工作場所審查及檢查辦法","Regulations for Review and Inspection of Dangerous Workplaces","民國109年7月17日","N0070019"),
 "OSH-32": ("製程安全評估定期實施辦法","Regulations for Periodic Process Safety Assessment","民國109年7月17日","N0060071"),
 "OSH-03": ("勞動檢查法","Labor Inspection Act","民國109年6月10日","N0070001"),
 "OSH-04": ("勞動檢查法施行細則","Enforcement Rules of the Labor Inspection Act","民國112年12月7日","N0070004"),
 "OSH-35": ("勞工職業災害保險及保護法","Labor Occupational Accident Insurance and Protection Act","民國110年4月30日","N0050031"),
 "OSH-22": ("女性勞工母性健康保護實施辦法","Regulations for Implementing Maternal Health Protection of Female Laborers","民國113年5月31日","N0060065"),
 "OSH-20": ("妊娠與分娩後女性及未滿十八歲勞工禁止從事危險性或有害性工作認定標準","Standards for Identifying Dangerous or Harmful Work Prohibited for Pregnant/Postpartum Women and Laborers under 18","民國114年11月20日","N0060032"),
 "OSH-21": ("事業單位僱用女性勞工夜間工作場所必要之安全衛生設施標準","Standards for Safety and Health Facilities for Female Laborers Working at Night","民國104年3月31日","N0030011"),
 "OSH-16": ("精密作業勞工視機能保護設施標準","Standards for Visual Protection Facilities for Precision Work","民國103年6月30日","N0060012"),
 "OSH-17": ("重體力勞動作業勞工保護措施標準","Standards for Protective Measures for Heavy Physical Labor","民國103年6月30日","N0060016"),
 "OSH-23": ("工業用機器人危害預防標準","Standards for Prevention of Industrial Robot Hazards","民國107年2月14日","N0060025"),
 "OSH-25": ("勞動檢查法第二十八條所定勞工有立即發生危險之虞認定標準","Standards for Identifying Imminent Danger under Article 28 of the Labor Inspection Act","民國94年6月10日","N0070016"),
 "OSH-29": ("職業安全衛生顧問服務機構與其顧問服務人員之認可及管理規則","Rules for Accreditation and Management of OSH Consulting Institutions","民國113年10月31日","N0060066"),
 "OSH-31": ("促進職業安全衛生文化獎勵及補助辦法","Regulations for Rewards and Subsidies to Promote OSH Culture","民國103年11月28日","N0060035"),
 "OSH-30": ("政府機關推動職業安全衛生業務績效評核及獎勵辦法","Regulations for Performance Evaluation and Rewards of Government OSH Promotion","民國103年11月27日","N0060061"),
 "OSH-19": ("辦理勞工體格與健康檢查醫療機構認可及管理辦法","Regulations for Accreditation and Management of Medical Institutions for Labor Health Examinations","民國111年8月12日","N0060040"),
 "OSH-27": ("違反職業安全衛生法及勞動檢查法案件處理要點","Directions for Handling Violations of the OSH Act and Labor Inspection Act","民國115年7月1日","mol:FL024857"),
 "OSH-26": ("推行職業安全衛生優良單位及人員選拔作業要點","Directions for Selection of Outstanding OSH Units and Personnel","民國114年1月17日","mol:FL015052"),
 "OSH-33": ("勞工體格與健康檢查特定檢查項目檢驗機構指定及管理作業要點","Directions for Designation of Laboratories for Specific Health Examination Items","民國112年5月22日","mol:FL078462"),
 "OSH-28": ("勞動部重大災害通報及檢查處理要點","MOL Directions for Major Accident Notification and Inspection","民國110年9月","isha:022"),
 "OSH-37": ("性別平等工作法","Act of Gender Equality in Employment","民國112年8月16日","N0030014"),
 "OSH-38": ("性別平等工作法施行細則","Enforcement Rules of the Act of Gender Equality in Employment","民國115年4月8日","N0030015"),
 "OSH-39": ("鉛中毒預防規則","Lead Poisoning Prevention Rules","","N0060018"),
 "OSH-40": ("四烷基鉛中毒預防規則","Tetraalkyl Lead Poisoning Prevention Rules","","N0060019"),
 "OSH-41": ("高壓氣體勞工安全規則","Rules for Labor Safety in High-Pressure Gas","","N0060030"),
 "OSH-42": ("起重升降機具安全規則","Safety Rules for Cranes, Hoists and Lifting Equipment","","N0060013"),
 "OSH-43": ("危害性化學品評估及分級管理辦法","Regulations for Assessment and Tiered Management of Hazardous Chemicals","","N0060070"),
 "OSH-44": ("鍋爐及壓力容器安全規則","Safety Rules for Boilers and Pressure Vessels","","N0060011"),
 "OSH-45": ("粉塵危害預防標準","Standards for Prevention of Dust Hazards","","N0060021"),
 "OSH-46": ("碼頭裝卸安全衛生設施標準","Safety and Health Facilities Standards for Wharf Loading and Unloading","","N0060006"),
 "OSH-47": ("有機溶劑中毒預防規則","Organic Solvent Poisoning Prevention Rules","","N0060017"),
 "OSH-48": ("勞動基準法","Labor Standards Act","","N0030001"),
 "OSH-49": ("勞動基準法施行細則","Enforcement Rules of the Labor Standards Act","","N0030002"),
 "OSH-50": ("特定化學物質危害預防標準","Standards for Prevention of Hazards from Specified Chemical Substances","","N0060015"),
}
LAW_VER = {k: (v[2] or law_version_from_text(v[0])) for k, v in LAWS.items()}

# 使用者指定納入遊戲之職安法規（2026-09-03）；其他職安法規之題目保留但 status=archived，Laws 權重 0
SCOPE_OSH = {"OSH-01","OSH-02","OSH-09","OSH-11","OSH-10","OSH-07","OSH-36","OSH-20","OSH-17","OSH-15",
             "OSH-14","OSH-16","OSH-03","OSH-04","OSH-25","OSH-23","OSH-22","OSH-08","OSH-12","OSH-34","OSH-18","OSH-37","OSH-38","OSH-39","OSH-40","OSH-41","OSH-42","OSH-43","OSH-44","OSH-45","OSH-46","OSH-47","OSH-48","OSH-49","OSH-50"}
# 使用者指定的 35 部職安法規（2026-09-04：「職安相關的只需要留我之前給的法條，其他移除」）
OSH_KEEP = {f"OSH-{i:02d}" for i in (1,2,3,4,7,8,9,10,11,12,14,15,16,17,18,20,22,23,25,34,36,37,38,39,40,41,42,43,44,45,46,47,48,49,50)}
def in_scope(lid):
    if lid.startswith("ENV"): return True
    return lid in OSH_KEEP

# Laws 分頁排序：依法規體系（母法→細則→子法）
FAMILIES = [
 ("職安-職業安全衛生法體系", [1,2,9,11,8]),
 ("職安-設施與作業標準", [7,36,34,14,15,16,17,18,23,46]),
 ("職安-危害性化學品與作業環境", [12,43,45,47,50,39,40]),
 ("職安-機械設備", [42,44,41]),
 ("職安-勞工健康與母性保護", [10,22,20]),
 ("職安-勞動檢查", [3,4,25]),
 ("職安-勞動條件與職場平等", [48,49,37,38]),
 ("環保-環境基本與環評", [21,13,63,64,65,66,98]),
 ("環保-空氣品質（含室內空品）", [1,2,26,27,28,29,32,33,34,35,3,4,36,37,38,30,31,99,100,101,102,103,104,105,106,107,108,110,111,112,113,114,115,116,117,118,119,120,121,122,123,124,125,126,127,128,129,131,132,133,134,135,136,137,139,140,141,142,143,144,145,146,147,148,149,150,151,152,153,208,209]),
 ("環保-噪音", [5,6,39,40,41,42,43,109,130,138,204,210]),
 ("環保-水污染與飲用水", [7,44,8,9,45,46,47,48,51,49,50,154,155,156,157,158,159,160,161,162,211,212]),
 ("環保-廢棄物與資源循環", [10,52,11,53,54,55,56,57,58,59,60,61,24,62,12,167,168,169,170,171,172,173,174,175,176,177,178,179,180,181,182,183,184,185,186,187]),
 ("環保-土壤及地下水", [19,77,78,79,80,81,82,83,201,202,203,222,223]),
 ("環保-毒性化學物質與環境用藥", [20,67,68,69,70,71,72,73,74,75,76,188,189,190,191,192,193,194,195,196,197,198,199,200,214,215,216,217,218,219,220,221]),
 ("環保-氣候變遷", [22,84,85,86,87,88,89,90,91,163,164,165,166,213]),
 ("環保-環境教育與環檢機構", [23,92,93,94,96,97,205,206,207]),
 ("環保-公害糾紛", [25,95]),
 ("環保-海洋與生態保育", [14,15,16,17,18]),
]
def family_order():
    out=[]
    for fam,ids in FAMILIES:
        g="OSH" if fam.startswith("職安") else "ENV"
        out += [(f"{g}-{i:02d}", fam) for i in ids]
    return out
LAW_NAME = {k: v[0] for k, v in LAWS.items()}

# 批次 → 模組清單（依序編號）
BATCHES = {
 1: ["batch1_a", "batch1_b", "batch1_c"],
 2: ["batch2_a", "batch2_b", "batch2_c", "batch2_d"],
 3: ["batch3_a", "batch3_b", "batch3_c"],
 4: ["batch4_a", "batch4_b", "batch4_c", "batch4_d"],
 5: ["batch5_a", "batch5_b", "batch5_c", "batch5_d"],
 6: ["batch6_a", "batch6_b", "batch6_c", "batch6_d", "batch6_e"],
 7: ["batch7_a", "batch7_b", "batch7_c", "batch7_d", "batch7_e", "batch7_f", "batch7_g"],
 8: ["batch8_a", "batch8_b", "batch8_c", "batch8_d"],
 9: sorted(_p.stem for _p in _Path(__file__).parent.glob("batch9_*.py")),
 10: sorted(_p.stem for _p in _Path(__file__).parent.glob("batch10_*.py")),
 11: sorted(_p.stem for _p in _Path(__file__).parent.glob("batch11_*.py")),
 12: sorted(_p.stem for _p in _Path(__file__).parent.glob("batch12_*.py")),
 13: sorted(_p.stem for _p in _Path(__file__).parent.glob("batch13_*.py")),
 14: sorted(_p.stem for _p in _Path(__file__).parent.glob("batch14_*.py")),
}
# 已歸檔（移到 _archive/）的題目檔自動略過
BATCHES = {k: [m for m in v if (_Path(__file__).parent / (m + ".py")).exists()] for k, v in BATCHES.items()}

# ---------- Laws 主檔（沿用批次 1 清單，另加 OSH-36 營造標準） ----------
OSH_LAWS = [
 ("職業安全衛生法","Occupational Safety and Health Act","N0060001",3),
 ("職業安全衛生法施行細則","Enforcement Rules of the OSH Act","N0060002",3),
 ("勞動檢查法","Labor Inspection Act","N0080001",2),
 ("勞動檢查法施行細則","Enforcement Rules of the Labor Inspection Act","N0080002",1),
 ("職業災害勞工保護法","Act for Protecting Workers of Occupational Accidents","N0060041",1),
 ("職業災害勞工保護法施行細則","Enforcement Rules of the Act for Protecting Workers of Occupational Accidents","N0060042",1),
 ("職業安全衛生設施規則","Occupational Safety and Health Facilities Rules","N0060009",3),
 ("危險性工作場所審查及檢查辦法","Regulations for Review and Inspection of Dangerous Workplaces","N0080006",2),
 ("職業安全衛生管理辦法","Regulations of Occupational Safety and Health Management","N0060003",3),
 ("勞工健康保護規則","Labor Health Protection Rules","N0060022",3),
 ("職業安全衛生教育訓練規則","Occupational Safety and Health Education and Training Rules","N0060008",3),
 ("危害性化學品標示及通識規則","Regulations on Labeling and Hazard Communication of Hazardous Chemicals","N0060052",2),
 ("機械設備器具安全標準","Safety Standards for Machinery, Equipment and Tools","N0060034",2),
 ("高溫作業勞工作息時間標準","Standards for Work and Rest Time of Laborers in High-Temperature Work","N0060019",2),
 ("高架作業勞工保護措施標準","Standards for Protective Measures for Laborers in Work at Height","N0060020",2),
 ("精密作業勞工視機能保護設施標準","Standards for Visual Protection Facilities for Precision Work","N0060012",1),
 ("重體力勞動作業勞工保護措施標準","Standards for Protective Measures for Heavy Physical Labor","N0060023",1),
 ("異常氣壓危害預防標準","Standards for Prevention of Abnormal Pressure Hazards","N0060004",2),
 ("辦理勞工體格與健康檢查醫療機構認可及管理辦法","Regulations for Accreditation and Management of Medical Institutions for Labor Health Examinations","N0060035",1),
 ("妊娠與分娩後女性及未滿十八歲勞工禁止從事危險性或有害性工作認定標準","Standards for Identifying Dangerous or Harmful Work Prohibited for Pregnant/Postpartum Women and Laborers under 18","N0060051",2),
 ("事業單位僱用女性勞工夜間工作場所必要之安全衛生設施標準","Standards for Safety and Health Facilities for Female Laborers Working at Night","N0060030",1),
 ("女性勞工母性健康保護實施辦法","Regulations for Implementing Maternal Health Protection of Female Laborers","N0060053",2),
 ("工業用機器人危害預防標準","Standards for Prevention of Industrial Robot Hazards","N0060007",1),
 ("職業安全衛生標示設置準則","Guidelines for Installation of Occupational Safety and Health Signs","N0060013",2),
 ("勞動檢查法第二十八條所定勞工有立即發生危險之虞認定標準","Standards for Identifying Imminent Danger under Article 28 of the Labor Inspection Act","N0080009",2),
 ("推行職業安全衛生優良單位及人員選拔作業要點","Directions for Selection of Outstanding OSH Units and Personnel","",1),
 ("違反職業安全衛生法及勞動檢查法案件處理要點","Directions for Handling Violations of the OSH Act and Labor Inspection Act","",1),
 ("勞動部重大災害通報及檢查處理要點","MOL Directions for Major Accident Notification and Inspection","",1),
 ("職業安全衛生顧問服務機構與其顧問服務人員之認可及管理規則","Rules for Accreditation and Management of OSH Consulting Institutions","N0060054",1),
 ("政府機關推動職業安全衛生業務績效評核及獎勵辦法","Regulations for Performance Evaluation and Rewards of Government OSH Promotion","N0060061",1),
 ("促進職業安全衛生文化獎勵及補助辦法","Regulations for Rewards and Subsidies to Promote OSH Culture","N0060055",1),
 ("製程安全評估定期實施辦法","Regulations for Periodic Process Safety Assessment","N0060050",1),
 ("勞工體格與健康檢查特定檢查項目檢驗機構指定及管理作業要點","Directions for Designation of Laboratories for Specific Health Examination Items","",1),
 ("缺氧症預防規則","Rules for Prevention of Hypoxia","N0060010",3),
 ("勞工職業災害保險及保護法","Labor Occupational Accident Insurance and Protection Act","N0060072",2),
 ("營造安全衛生設施標準","Construction Safety and Health Facilities Standards","N0060014",3),
 ("性別平等工作法","Act of Gender Equality in Employment","N0030014",1),
 ("性別平等工作法施行細則","Enforcement Rules of the Act of Gender Equality in Employment","N0030015",1),
 ("鉛中毒預防規則","Lead Poisoning Prevention Rules","N0060018",1),
 ("四烷基鉛中毒預防規則","Tetraalkyl Lead Poisoning Prevention Rules","N0060019",1),
 ("高壓氣體勞工安全規則","Rules for Labor Safety in High-Pressure Gas","N0060030",1),
 ("起重升降機具安全規則","Safety Rules for Cranes, Hoists and Lifting Equipment","N0060013",1),
 ("危害性化學品評估及分級管理辦法","Regulations for Assessment and Tiered Management of Hazardous Chemicals","N0060070",1),
 ("鍋爐及壓力容器安全規則","Safety Rules for Boilers and Pressure Vessels","N0060011",1),
 ("粉塵危害預防標準","Standards for Prevention of Dust Hazards","N0060021",1),
 ("碼頭裝卸安全衛生設施標準","Safety and Health Facilities Standards for Wharf Loading and Unloading","N0060006",1),
 ("有機溶劑中毒預防規則","Organic Solvent Poisoning Prevention Rules","N0060017",1),
 ("勞動基準法","Labor Standards Act","N0030001",1),
 ("勞動基準法施行細則","Enforcement Rules of the Labor Standards Act","N0030002",1),
 ("特定化學物質危害預防標準","Standards for Prevention of Hazards from Specified Chemical Substances","N0060015",1),
 ("勞工作業場所容許暴露標準","Standards of Permissible Exposure Limits at Job Sites","N0060004",3),
 ("勞工作業環境監測實施辦法","Regulations Governing Workplace Environment Monitoring","N0060033",3),
 ("危險性機械及設備安全檢查規則","Safety Inspection Rules for Dangerous Machinery and Equipment","N0060039",2),
 ("既有危險性機械及設備安全檢查規則","Safety Inspection Rules for Existing Dangerous Machinery and Equipment","N0060053",1),
 ("職場霸凌防治措施準則","Guidelines for Workplace Bullying Prevention Measures","N0060085",3),
 ("地方主管機關受理最高負責人職場霸凌事件申訴處理辦法","Regulations for Local Authorities Handling Bullying Complaints against Top Executives","N0060086",1),
 ("工程安全設計及整體工程統合管理辦法","Regulations on Engineering Safety Design and Integrated Project Management","N0060088",2),
 ("新化學物質登記管理辦法","Regulations on Registration of New Chemical Substances","N0060069",1),
 ("管制性化學品之指定及運作許可管理辦法","Regulations on Designation and Operating Permits of Controlled Chemicals","N0060068",1),
 ("優先管理化學品之指定及運作管理辦法","Regulations on Designation and Operation Management of Priority Management Chemicals","N0060064",1),
 ("機械設備器具安全資訊申報登錄辦法","Regulations on Safety Information Registration of Machinery, Equipment and Tools","N0060056",1),
 ("機械設備器具監督管理辦法","Regulations on Supervision and Management of Machinery, Equipment and Tools","N0060063",1),
 ("機械類產品型式驗證實施及監督管理辦法","Regulations on Type Certification of Machinery Products","N0060062",1),
 ("機械類產品申請先行放行辦法","Regulations on Advance Release of Machinery Products","N0060057",1),
 ("機械類產品申請免驗證辦法","Regulations on Exemption from Certification of Machinery Products","N0060059",1),
 ("構造規格特殊產品安全評估報告及檢驗辦法","Regulations on Safety Assessment Reports for Products of Special Construction","N0060060",1),
 ("安全標示與驗證合格標章使用及管理辦法","Regulations on Use of Safety Labels and Certification Marks","N0060058",1),
 ("固定式起重機安全檢查構造標準","Construction Standards for Safety Inspection of Fixed Cranes","N0070022",1),
 ("移動式起重機安全檢查構造標準","Construction Standards for Safety Inspection of Mobile Cranes","N0070021",1),
 ("升降機安全檢查構造標準","Construction Standards for Safety Inspection of Lifts","N0070017",1),
 ("吊籠安全檢查構造標準","Construction Standards for Safety Inspection of Gondolas","N0070023",1),
 ("壓力容器安全檢查構造標準","Construction Standards for Safety Inspection of Pressure Vessels","N0060055",1),
 ("林場安全衛生設施規則","Forestry Safety and Health Facilities Rules","N0060005",1),
 ("礦場職業衛生設施標準","Mine Occupational Health Facilities Standards","N0060003",1),
 ("船舶清艙解體職業安全規則","Occupational Safety Rules for Ship Tank Cleaning and Breaking","N0060031",1),
 ("勞工健康服務專業機構管理規則","Rules for Management of Labor Health Service Institutions","N0060087",1),
 ("勞工職業災害保險預防職業病健康檢查及健康追蹤檢查辦法","Regulations on Preventive Occupational Disease Health Examinations under Occupational Accident Insurance","N0060077",1),
 ("勞工職業災害保險職業病鑑定作業實施辦法","Regulations on Occupational Disease Determination under Occupational Accident Insurance","N0060080",1),
 ("職業災害勞工補助及核發辦法","Regulations on Subsidies for Workers with Occupational Accidents","N0060073",1),
 ("職業災害勞工申請器具照護失能及死亡補助辦法","Regulations on Aids, Care, Disability and Death Subsidies for Occupational Accident Workers","N0060076",1),
 ("職業災害勞工職業重建補助辦法","Regulations on Vocational Rehabilitation Subsidies for Occupational Accident Workers","N0060075",1),
 ("職業災害勞工職能復健專業機構認可管理及補助辦法","Regulations on Accreditation of Occupational Rehabilitation Institutions","N0060082",1),
 ("職業災害預防及職業災害勞工重建補助辦法","Regulations on Subsidies for Occupational Accident Prevention and Worker Rehabilitation","N0060081",1),
 ("職業災害預防補助辦法","Regulations on Occupational Accident Prevention Subsidies","N0060049",1),
 ("職業傷病診治醫療機構認可管理補助及職業傷病通報辦法","Regulations on Accreditation of Occupational Injury and Disease Medical Institutions and Notification","N0060083",1),
 ("直轄市及縣市政府辦理協助職業災害勞工重返職場補助辦法","Regulations on Local Government Subsidies for Return-to-Work of Occupational Accident Workers","N0060079",1),
 ("財團法人職業災害預防及重建中心監督及管理辦法","Regulations on Supervision of the Occupational Accident Prevention and Rehabilitation Center","N0060078",1),
 ("危險性機械或設備代行檢查機構管理規則","Rules for Management of Designated Inspection Agencies for Dangerous Machinery and Equipment","N0070018",1),
 ("勞動檢查員遴用及專業訓練辦法","Regulations on Recruitment and Training of Labor Inspectors","N0070006",1),
 ("勞動檢查員執行職務迴避辦法","Regulations on Recusal of Labor Inspectors","N0070005",1),
]
ENV_LAWS = [
 ("空氣污染防制法","Air Pollution Control Act","O0020001",3,1),
 ("空氣污染防制法施行細則","Enforcement Rules of the Air Pollution Control Act","O0020002",1,1),
 ("營建工程空氣污染防制設施管理辦法","Regulations on Air Pollution Control Facilities for Construction Projects","O0020021",3,1),
 ("固定污染源逸散性粒狀污染物空氣污染防制設施管理辦法","Regulations on Control Facilities for Fugitive Particulate Pollutants from Stationary Sources","O0020033",2,1),
 ("噪音管制法","Noise Control Act","O0030001",2,1),
 ("噪音管制標準","Noise Control Standards","O0030004",2,1),
 ("水污染防治法","Water Pollution Control Act","O0040001",2,1),
 ("放流水標準","Effluent Standards","O0040004",1,1),
 ("水污染防治措施及檢測申報管理辦法","Regulations on Water Pollution Control Measures and Monitoring Reporting","O0040030",1,1),
 ("廢棄物清理法","Waste Disposal Act","O0050001",3,1),
 ("事業廢棄物貯存清除處理方法及設施標準","Standards for Storage, Clearance and Disposal of Industrial Waste","O0050023",2,1),
 ("營建剩餘土石方處理方案","Construction Surplus Soil and Rock Disposal Program","",2,1),
 ("環境影響評估法","Environmental Impact Assessment Act","O0090001",2,1),
 ("海洋污染防治法","Marine Pollution Control Act","O0040026",2,2),
 ("海岸管理法","Coastal Zone Management Act","D0070212",1,2),
 ("海洋保育法","Marine Conservation Act","",1,2),
 ("野生動物保育法","Wildlife Conservation Act","M0120001",2,2),
 ("濕地保育法","Wetland Conservation Act","D0130001",1,2),
 ("土壤及地下水污染整治法","Soil and Groundwater Pollution Remediation Act","O0110001",1,2),
 ("毒性及關注化學物質管理法","Toxic and Concerned Chemical Substances Control Act","O0060001",1,2),
 ("環境基本法","Basic Environment Act","O0100001",1,3),
 ("氣候變遷因應法","Climate Change Response Act","O0020098",1,3),
 ("環境教育法","Environmental Education Act","O0120001",1,3),
 ("資源循環推動法","Resource Circulation Promotion Act","O0050049",1,3),
 ("公害糾紛處理法","Public Nuisance Dispute Mediation Act","O0100002",1,3),
 ('空氣污染防制專責單位或專責人員設置及管理辦法','Regulations on Air Pollution Control Dedicated Units and Personnel','O0020106',2,1),
 ('固定污染源空氣污染物排放標準','Stationary Source Air Pollutant Emission Standards','O0020006',2,1),
 ('空氣品質標準','Air Quality Standards','O0020007',2,1),
 ('空氣品質嚴重惡化警告發布及緊急防制辦法','Regulations on Severe Air Quality Deterioration Warnings and Emergency Control','O0020015',1,1),
 ('室內空氣品質管理法','Indoor Air Quality Management Act','O0130001',2,1),
 ('室內空氣品質標準','Indoor Air Quality Standards','O0130005',1,1),
 ('固定污染源設置操作及燃料使用許可證管理辦法','Regulations on Stationary Source Installation, Operation and Fuel Use Permits','O0020012',1,1),
 ('空氣污染防制費收費辦法','Regulations on Air Pollution Control Fees','O0020027',1,1),
 ('揮發性有機物空氣污染管制及排放標準','VOC Air Pollution Control and Emission Standards','O0020030',1,1),
 ('鍋爐空氣污染物排放標準','Boiler Air Pollutant Emission Standards','O0020113',1,1),
 ('汽車停車怠速管理辦法','Regulations on Vehicle Idling Management','O0020086',1,1),
 ('空氣污染行為管制執行準則','Guidelines for Enforcement of Air Pollution Behavior Control','O0020036',1,1),
 ('空氣品質監測站設置及監測準則','Guidelines for Air Quality Monitoring Station Siting and Monitoring','O0020121',1,1),
 ('噪音管制法施行細則','Enforcement Rules of the Noise Control Act','O0030002',1,1),
 ('環境音量標準','Environmental Noise Standards','O0030014',2,1),
 ('噪音管制區劃定作業準則','Guidelines for Designating Noise Control Zones','O0030016',1,1),
 ('易發生噪音設施設置及操作許可辦法','Regulations on Permits for Noise-Prone Facilities','O0030007',1,1),
 ('機動車輛噪音管制標準','Motor Vehicle Noise Control Standards','O0030013',1,1),
 ('水污染防治法施行細則','Enforcement Rules of the Water Pollution Control Act','O0040002',1,1),
 ('地面水體分類及水質標準','Surface Water Classification and Quality Standards','O0040005',1,1),
 ('廢（污）水處理專責單位或人員設置及管理辦法','Regulations on Wastewater Treatment Dedicated Units and Personnel','O0040070',2,1),
 ('水污染防治措施計畫及許可申請審查管理辦法','Regulations on Water Pollution Control Plans and Permit Review','O0040055',1,1),
 ('事業或污水下水道系統排放廢（污）水緊急應變辦法','Regulations on Emergency Response for Wastewater Discharge','O0040049',1,1),
 ('飲用水管理條例','Drinking Water Management Act','O0040010',1,1),
 ('飲用水水質標準','Drinking Water Quality Standards','O0040019',1,1),
 ('土壤處理標準','Soil Treatment Standards','O0040031',1,1),
 ('廢棄物清理法施行細則','Enforcement Rules of the Waste Disposal Act','O0050036',1,1),
 ('有害事業廢棄物認定標準','Standards for Defining Hazardous Industrial Waste','O0050023',2,1),
 ('公民營廢棄物清除處理機構許可管理辦法','Regulations on Permits for Public and Private Waste Clearance and Disposal Organizations','O0050039',2,1),
 ('事業廢棄物清理計畫書審查管理辦法','Regulations on Review of Industrial Waste Disposal Plans','O0050084',1,1),
 ('廢棄物清理專業技術人員管理辦法','Regulations on Waste Disposal Professional Technicians','O0050053',1,1),
 ('一般廢棄物回收清除處理辦法','Regulations on General Waste Recycling, Clearance and Disposal','O0050024',1,1),
 ('應回收廢棄物責任業者管理辦法','Regulations on Responsible Enterprises for Recyclable Waste','O0050062',1,1),
 ('事業委託清理之相當注意義務認定準則','Guidelines on Due Care in Entrusting Waste Disposal','O0050085',1,1),
 ('環境部事業廢棄物再利用管理辦法','MOENV Regulations on Industrial Waste Reuse','O0050082',1,1),
 ('共通性事業廢棄物再利用管理辦法','Regulations on Reuse of Common Industrial Waste','O0050086',1,1),
 ('資源回收再利用法施行細則','Enforcement Rules of the Resource Recycling Act','O0050076',1,1),
 ('環境影響評估法施行細則','Enforcement Rules of the Environmental Impact Assessment Act','O0090002',1,1),
 ('開發行為應實施環境影響評估細目及範圍認定標準','Standards for Determining Development Activities Requiring EIA','O0090012',2,1),
 ('開發行為環境影響評估作業準則','Guidelines for EIA of Development Activities','O0090003',1,1),
 ('政府政策環境影響評估作業辦法','Regulations on Strategic Environmental Assessment of Government Policies','O0090029',1,1),
 ('毒性及關注化學物質管理法施行細則','Enforcement Rules of the Toxic and Concerned Chemical Substances Control Act','O0060013',1,2),
 ('毒性及關注化學物質危害預防及應變計畫作業辦法','Regulations on Hazard Prevention and Response Plans for Toxic and Concerned Chemicals','O0060035',2,2),
 ('毒性及關注化學物質專業技術管理人員設置及管理辦法','Regulations on Professional Technical Managers for Toxic and Concerned Chemicals','O0060046',2,2),
 ('毒性及關注化學物質標示與安全資料表管理辦法','Regulations on Labeling and SDS for Toxic and Concerned Chemicals','O0060037',1,2),
 ('毒性及關注化學物質運送管理辦法','Regulations on Transport of Toxic and Concerned Chemicals','O0060015',1,2),
 ('毒性及關注化學物質許可登記核可管理辦法','Regulations on Permits, Registration and Approval of Toxic and Concerned Chemicals','O0060038',1,2),
 ('毒性及關注化學物質運作與釋放量紀錄管理辦法','Regulations on Operation and Release Records of Toxic and Concerned Chemicals','O0060039',1,2),
 ('毒性及關注化學物質應變器材與偵測警報設備管理辦法','Regulations on Response Equipment and Detection Alarms for Toxic and Concerned Chemicals','O0060036',1,2),
 ('新化學物質及既有化學物質資料登錄辦法','Regulations on Registration of New and Existing Chemical Substances','O0060043',1,2),
 ('環境用藥管理法','Environmental Agents Control Act','O0060001',1,2),
 ('土壤及地下水污染整治法施行細則','Enforcement Rules of the Soil and Groundwater Pollution Remediation Act','O0110002',1,2),
 ('土壤污染管制標準','Soil Pollution Control Standards','O0110005',2,2),
 ('地下水污染管制標準','Groundwater Pollution Control Standards','O0110006',2,2),
 ('土壤污染監測標準','Soil Pollution Monitoring Standards','O0110012',1,2),
 ('地下水污染監測標準','Groundwater Pollution Monitoring Standards','O0110013',1,2),
 ('土壤及地下水污染場址初步評估暨處理等級評定辦法','Regulations on Preliminary Assessment and Grading of Contaminated Sites','O0110022',1,2),
 ('防止貯存系統污染地下水體設施及監測設備設置管理辦法','Regulations on Storage System Groundwater Protection Facilities and Monitoring','O0110010',1,2),
 ('氣候變遷因應法施行細則','Enforcement Rules of the Climate Change Response Act','O0020103',1,3),
 ('溫室氣體排放量盤查登錄及查驗管理辦法','Regulations on GHG Emission Inventory, Registration and Verification','O0020102',2,3),
 ('碳費收費辦法','Carbon Fee Collection Regulations','O0020139',2,3),
 ('自主減量計畫管理辦法','Regulations on Voluntary Reduction Plans','O0020140',1,3),
 ('溫室氣體減量額度交易拍賣及移轉管理辦法','Regulations on Trading, Auction and Transfer of GHG Reduction Credits','O0020138',1,3),
 ('溫室氣體自願減量專案管理辦法','Regulations on Voluntary GHG Reduction Projects','O0020137',1,3),
 ('溫室氣體排放量增量抵換管理辦法','Regulations on Offsetting Incremental GHG Emissions','O0020136',1,3),
 ('自願性產品碳足跡核定標示及管理辦法','Regulations on Voluntary Product Carbon Footprint Labeling','O0020142',1,3),
 ('環境教育法施行細則','Enforcement Rules of the Environmental Education Act','O0120002',1,3),
 ('環境講習執行辦法','Regulations on Environmental Lectures','O0120010',1,3),
 ('環境教育人員認證及管理辦法','Regulations on Certification of Environmental Education Personnel','O0120006',1,3),
 ('公害糾紛處理法施行細則','Enforcement Rules of the Public Nuisance Dispute Mediation Act','O0080002',1,3),
 ('環境保護專責及技術人員訓練管理辦法','Regulations on Training of Environmental Protection Dedicated and Technical Personnel','O0100006',2,3),
 ('環境檢驗測定機構管理辦法','Regulations on Environmental Testing Organizations','O0070001',1,3),
 ("違反環境影響評估法按日連續處罰執行準則","Guidelines for the Execution of Consecutive Daily Fines for Violations of the Environmental Impact Assessment Act","O0090030",1,2),
 ("中小型廢棄物焚化爐戴奧辛管制及排放標準","Dioxin Control and Emission Standards for Small and Medium-sized Waste Incinerators","O0020037",1,2),
 ("公告場所室內空氣品質檢驗測定管理辦法","Regulations Governing Indoor Air Quality Testing and Measurement of Announced Premises","O0130006",1,2),
 ("公私場所固定污染源申請改善排放空氣污染物總量及濃度管理辦法","Regulations Governing Public and Private Premises' Applications for Improving the Total Quantity and Concentration of Air Pollutants Emitted from Stationary Pollution Sources","O0020061",1,2),
 ("公私場所固定污染源空氣污染物排放量申報管理辦法","Regulations Governing the Reporting of Air Pollutant Emissions from Stationary Pollution Sources of Public and Private Premises","O0020066",1,2),
 ("公私場所固定污染源復工試車評鑑及管理辦法","Regulations Governing the Resumption of Operation, Trial Run Evaluation, and Management of Stationary Pollution Sources of Public and Private Premises","O0020063",1,2),
 ("公私場所固定污染源違反空氣污染防制法應處罰鍰額度裁罰準則","Guidelines for Determining Fine Amounts for Stationary Pollution Sources of Public and Private Premises in Violation of the Air Pollution Control Act","O0020038",1,2),
 ("公私場所固定污染源燃料混燒比例成分及防制設施管制標準","Control Standards for Fuel Co-firing Ratio, Composition, and Control Equipment of Stationary Pollution Sources of Public and Private Premises","O0020124",1,2),
 ("水泥業空氣污染物排放標準","Air Pollutant Emission Standards for the Cement Industry","O0020041",1,2),
 ("加油站油氣回收設施管理辦法","Regulations Governing Vapor Recovery Facilities at Gas Stations","O0020045",1,2),
 ("半導體製造業空氣污染管制及排放標準","Air Pollution Control and Emission Standards for the Semiconductor Manufacturing Industry","O0020032",1,2),
 ("民用航空器噪音管制標準","Noise Control Standards for Civil Aircraft","O0030004",1,2),
 ("光電材料及元件製造業空氣污染管制及排放標準","Air Pollution Control and Emission Standards for the Optoelectronic Materials and Components Manufacturing Industry","O0020073",1,2),
 ("汽車製造業表面塗裝作業空氣污染物排放標準","Air Pollutant Emission Standards for Surface Coating Operations of the Motor Vehicle Manufacturing Industry","O0020029",1,2),
 ("汽油及替代清潔燃料引擎汽車排放空氣污染物檢驗站設置及管理辦法","Regulations Governing the Establishment and Management of Inspection Stations for Air Pollutant Emissions from Gasoline and Alternative Clean Fuel Engine Vehicles","O0020135",1,2),
 ("固定污染源有害空氣污染物排放標準","Hazardous Air Pollutant Emission Standards for Stationary Pollution Sources","O0020128",1,2),
 ("固定污染源自行或委託檢測及申報管理辦法","Regulations Governing Self or Commissioned Testing and Reporting by Stationary Pollution Sources","O0020054",1,2),
 ("固定污染源空氣污染物連續自動監測設施管理辦法","Regulations Governing Continuous Automated Monitoring Facilities for Air Pollutants from Stationary Pollution Sources","O0020068",1,2),
 ("固定污染源戴奧辛排放標準","Dioxin Emission Standards for Stationary Pollution Sources","O0020072",1,2),
 ("易致空氣污染之物質使用許可證管理辦法","Regulations Governing Permits for the Use of Substances Prone to Cause Air Pollution","O0020059",1,2),
 ("室內空氣品質管理法施行細則","Enforcement Rules of the Indoor Air Quality Management Act","O0130002",1,2),
 ("室內空氣品質維護管理專責人員設置管理辦法","Regulations Governing the Assignment and Management of Dedicated Personnel for Indoor Air Quality Maintenance and Management","O0130003",1,2),
 ("建物及工業維護塗料揮發性有機物成分標準","Volatile Organic Compound Content Standards for Architectural and Industrial Maintenance Coatings","O0020119",1,2),
 ("氟氯烴管理辦法","Regulations Governing the Management of Chlorofluorocarbons","O0020052",1,2),
 ("玻璃業空氣污染物排放標準","Air Pollutant Emission Standards for the Glass Industry","O0020018",1,2),
 ("特殊性工業區緩衝地帶及空氣品質監測設施設置標準","Standards for the Establishment of Buffer Zones and Air Quality Monitoring Facilities in Special Industrial Zones","O0020022",1,2),
 ("乾洗作業空氣污染防制設施管制標準","Control Standards for Air Pollution Control Equipment in Dry Cleaning Operations","O0020033",1,2),
 ("移動污染源空氣污染防制設備管理辦法","Regulations Governing Air Pollution Control Equipment for Mobile Pollution Sources","O0020025",1,2),
 ("移動污染源空氣污染物排放標準","Air Pollutant Emission Standards for Mobile Pollution Sources","O0020003",1,2),
 ("移動污染源違反空氣污染防制法裁罰準則","Guidelines for Fines on Mobile Pollution Sources in Violation of the Air Pollution Control Act","O0020043",1,2),
 ("移動污染源燃料成分管制標準","Fuel Composition Control Standards for Mobile Pollution Sources","O0020034",1,2),
 ("移動污染源燃料販賣進口許可及管理辦法","Regulations Governing the Permitting and Management of the Sale and Import of Fuel for Mobile Pollution Sources","O0020065",1,2),
 ("陸上運輸系統噪音管制標準","Noise Control Standards for Land Transportation Systems","O0030018",1,2),
 ("陶瓷業空氣污染物排放標準","Air Pollutant Emission Standards for the Ceramics Industry","O0020019",1,2),
 ("氯乙烯及聚氯乙烯製造業空氣污染物管制及排放標準","Air Pollution Control and Emission Standards for the Vinyl Chloride and Polyvinyl Chloride Manufacturing Industry","O0020115",1,2),
 ("溴化甲烷管理辦法","Regulations Governing the Management of Methyl Bromide","O0020057",1,2),
 ("煉鋼及鑄造電爐粒狀污染物管制及排放標準","Particulate Matter Control and Emission Standards for Steelmaking and Foundry Electric Arc Furnaces","O0020017",1,2),
 ("煉鋼業電弧爐戴奧辛管制及排放標準","Dioxin Control and Emission Standards for Electric Arc Furnaces in the Steelmaking Industry","O0020044",1,2),
 ("違反空氣污染防制法按次處罰通知限期改善補正或申報執行準則","Guidelines for the Execution of Per-Violation Fines and Notices to Improve, Correct, or Report within a Specified Period for Violations of the Air Pollution Control Act","O0020062",1,2),
 ("違反室內空氣品質管理法罰鍰額度裁罰準則","Guidelines for Determining Fine Amounts for Violations of the Indoor Air Quality Management Act","O0130004",1,2),
 ("違反噪音管制法按日連續處罰執行準則","Guidelines for the Execution of Consecutive Daily Fines for Violations of the Noise Control Act","O0030019",1,2),
 ("鉛二次冶煉廠空氣污染物排放標準","Air Pollutant Emission Standards for Secondary Lead Smelters","O0020008",1,2),
 ("電力設施空氣污染物排放標準","Air Pollutant Emission Standards for Power Generation Facilities","O0020026",1,2),
 ("聚氨基甲酸酯塗布業揮發性有機物空氣污染管制及排放標準","Volatile Organic Compound Air Pollution Control and Emission Standards for the Polyurethane Coating Industry","O0020071",1,2),
 ("蒙特婁議定書列管化學物質管理辦法","Regulations Governing the Management of Chemical Substances Controlled under the Montreal Protocol","O0020074",1,2),
 ("廢棄物焚化爐空氣污染物排放標準","Air Pollutant Emission Standards for Waste Incinerators","O0020010",1,2),
 ("廢棄物焚化爐戴奧辛管制及排放標準","Dioxin Control and Emission Standards for Waste Incinerators","O0020031",1,2),
 ("熱風乾燥機粒狀污染物排放標準","Particulate Matter Emission Standards for Hot Air Dryers","O0020028",1,2),
 ("膠帶製造業揮發性有機物空氣污染管制及排放標準","Volatile Organic Compound Air Pollution Control and Emission Standards for the Adhesive Tape Manufacturing Industry","O0020077",1,2),
 ("機車排放空氣污染物檢驗站設置及管理辦法","Regulations Governing the Establishment and Management of Inspection Stations for Air Pollutant Emissions from Motorcycles","O0020056",1,2),
 ("磚瓦窯業開放式隧道窯粒狀污染物排放標準","Particulate Matter Emission Standards for Open Tunnel Kilns in the Brick and Tile Industry","O0020020",1,2),
 ("鋼鐵業集塵灰高溫冶煉設施戴奧辛管制及排放標準","Dioxin Control and Emission Standards for High-temperature Smelting Facilities for Collected Dust in the Steel Industry","O0020070",1,2),
 ("鋼鐵業燒結工場空氣污染物排放標準","Air Pollutant Emission Standards for Sintering Plants in the Steel Industry","O0020011",1,2),
 ("鋼鐵業燒結工場戴奧辛管制及排放標準","Dioxin Control and Emission Standards for Sintering Plants in the Steel Industry","O0020069",1,2),
 ("餐飲業空氣污染防制設施管理辦法","Regulations Governing Air Pollution Control Equipment for the Food and Beverage Service Industry","O0020127",1,2),
 ("瀝青拌合業粒狀污染物排放標準","Particulate Matter Emission Standards for the Asphalt Mixing Industry","O0020021",1,2),
 ("海洋放流管線放流水標準","Effluent Standards for Ocean Outfall Pipelines","O0040013",1,2),
 ("飲用水水源水質標準","Water Quality Standards for Drinking Water Sources","O0040018",1,2),
 ("飲用水連續供水固定設備使用及維護管理辦法","Regulations Governing the Use and Maintenance of Fixed Equipment for Continuous Drinking Water Supply","O0040014",1,2),
 ("飲用水管理條例施行細則","Enforcement Rules of the Drinking Water Management Act","O0040016",1,2),
 ("違反水污染防治法按次處罰通知限期改善或補正執行準則","Guidelines for the Execution of Per-Violation Fines and Notices to Improve or Correct within a Specified Period for Violations of the Water Pollution Control Act","O0040065",1,2),
 ("違反水污染防治法罰鍰額度裁罰準則","Guidelines for Determining Fine Amounts for Violations of the Water Pollution Control Act","O0040056",1,2),
 ("違反飲用水管理條例按日連續處罰執行準則","Guidelines for the Execution of Consecutive Daily Fines for Violations of the Drinking Water Management Act","O0040052",1,2),
 ("預鑄式建築物污水處理設施管理辦法","Regulations Governing Sewage Treatment Facilities for Prefabricated Buildings","O0040047",1,2),
 ("廢（污）水處理專責人員違反水污染防治法罰鍰額度裁罰準則","Guidelines for Determining Fine Amounts for Wastewater Treatment Dedicated Personnel in Violation of the Water Pollution Control Act","O0040069",1,2),
 ("二氧化碳捕捉後封存管理辦法","Regulations Governing the Management of Carbon Dioxide Capture and Storage","O0020145",1,2),
 ("氫氟碳化物管理辦法","Regulations Governing the Management of Hydrofluorocarbons","O0020141",1,2),
 ("溫室氣體抵換專案管理辦法","Regulations Governing the Management of Greenhouse Gas Offset Projects","O0020101",1,2),
 ("溫室氣體認證機構及查驗機構管理辦法","Regulations Governing the Management of Greenhouse Gas Certification Bodies and Verification Bodies","O0020104",1,2),
 ("共通性事業廢棄物作為固體再生燃料原料再利用管理辦法","Regulations Governing the Reuse of Common Industrial Waste as Raw Material for Solid Recovered Fuel","O0050091",1,2),
 ("再生資源再使用管理辦法","Regulations Governing the Reuse of Recycled Resources","O0050073",1,2),
 ("再生資源限制或禁止輸入輸出管理辦法","Regulations Governing the Restriction or Prohibition of Import and Export of Recycled Resources","O0050074",1,2),
 ("有害事業廢棄物檢測及紀錄管理辦法","Regulations Governing the Testing and Recording of Hazardous Industrial Waste","O0070006",1,2),
 ("事業自行清除處理事業廢棄物許可管理辦法","Regulations Governing Permits for Enterprises to Self-Clear and Dispose of Industrial Waste","O0050069",1,2),
 ("事業廢棄物處理設施餘裕處理容量許可管理辦法","Regulations Governing Permits for Surplus Treatment Capacity of Industrial Waste Treatment Facilities","O0050045",1,2),
 ("事業廢棄物輸入輸出管理辦法","Regulations Governing the Import and Export of Industrial Waste","O0050016",1,2),
 ("依促進民間參與公共建設法設置之廢棄物清除處理設施管理辦法","Regulations Governing Waste Clearance and Disposal Facilities Established under the Act for Promotion of Private Participation in Infrastructure Projects","O0050047",1,2),
 ("指定公營事業設置廢棄物清除處理設施管理辦法","Regulations Governing the Establishment of Waste Clearance and Disposal Facilities by Designated State-run Enterprises","O0050050",1,2),
 ("違反廢棄物清理法按日連續處罰執行準則","Guidelines for the Execution of Consecutive Daily Fines for Violations of the Waste Disposal Act","O0050043",1,2),
 ("違反廢棄物清理法罰鍰額度裁罰準則","Guidelines for Determining Fine Amounts for Violations of the Waste Disposal Act","O0050090",1,2),
 ("廢容器回收貯存清除處理方法及設施標準","Standards for the Recycling, Storage, Clearance, and Disposal Methods and Facilities for Waste Containers","O0050054",1,2),
 ("廢乾電池回收貯存清除處理方法及設施標準","Standards for the Recycling, Storage, Clearance, and Disposal Methods and Facilities for Waste Dry-cell Batteries","O0050059",1,2),
 ("廢照明光源回收貯存清除處理方法及設施標準","Standards for the Recycling, Storage, Clearance, and Disposal Methods and Facilities for Waste Lighting Sources","O0050056",1,2),
 ("廢鉛蓄電池回收貯存清除處理方法及設施標準","Standards for the Recycling, Storage, Clearance, and Disposal Methods and Facilities for Waste Lead-acid Batteries","O0050061",1,2),
 ("廢電子電器暨廢資訊物品回收貯存清除處理方法及設施標準","Standards for the Recycling, Storage, Clearance, and Disposal Methods and Facilities for Waste Electronic and Electrical Equipment and Waste Information Products","O0050079",1,2),
 ("廢輪胎回收貯存清除處理方法及設施標準","Standards for the Recycling, Storage, Clearance, and Disposal Methods and Facilities for Waste Tires","O0050060",1,2),
 ("廢機動車輛回收貯存清除處理方法及設施標準","Standards for the Recycling, Storage, Clearance, and Disposal Methods and Facilities for Waste Motor Vehicles","O0050064",1,2),
 ("應回收廢棄物回收清除處理補貼申請審核管理辦法","Regulations Governing the Application and Review of Subsidies for the Recycling, Clearance, and Disposal of Mandatory Recyclable Waste","O0050057",1,2),
 ("應回收廢棄物回收處理業管理辦法","Regulations Governing Recycling and Disposal Enterprises of Mandatory Recyclable Waste","O0050051",1,2),
 ("環境部再生資源再生利用管理辦法","Regulations Governing the Recycling and Reuse of Recycled Resources by the Ministry of Environment","O0050072",1,2),
 ("毒性及關注化學物質災害事故應變車輛管理辦法","Regulations Governing Emergency Response Vehicles for Disasters and Accidents Involving Toxic and Concerned Chemical Substances","O0060045",1,2),
 ("毒性及關注化學物質災害與懸浮微粒物質災害救助種類及標準","Types and Standards for Disaster Relief for Disasters Involving Toxic and Concerned Chemical Substances and Suspended Particulate Matter","O0060017",1,2),
 ("毒性及關注化學物質專業應變人員管理辦法","Regulations Governing Professional Emergency Response Personnel for Toxic and Concerned Chemical Substances","O0060053",1,2),
 ("毒性及關注化學物質環境事故專業應變諮詢機關（構）認證及管理辦法","Regulations Governing the Accreditation and Management of Professional Emergency Response Consultation Agencies (Institutions) for Environmental Incidents Involving Toxic and Concerned Chemical Substances","O0060051",1,2),
 ("病媒防治業管理辦法","Regulations Governing the Vector Control Industry","O0060011",1,2),
 ("違反毒性及關注化學物質管理法罰鍰額度裁罰準則","Guidelines for Determining Fine Amounts for Violations of the Toxic and Concerned Chemical Substances Control Act","O0060050",1,2),
 ("遺傳工程環境用藥微生物製劑開發試驗研究管理辦法","Regulations Governing the Development, Testing, and Research of Genetically Engineered Environmental Pesticide Microbial Preparations","O0060009",1,2),
 ("環境用藥工廠設廠標準","Standards for the Establishment of Environmental Pesticide Factories","O0060010",1,2),
 ("環境用藥專業技術人員設置管理辦法","Regulations Governing the Assignment and Management of Professional Technical Personnel for Environmental Pesticides","O0060007",1,2),
 ("環境用藥貯存置放使用管理辦法","Regulations Governing the Storage, Placement, and Use of Environmental Pesticides","O0060021",1,2),
 ("環境用藥微生物製劑使用於生態及水源保育或保護區運作管理辦法","Regulations Governing the Use of Microbial Pesticide Preparations in Ecological and Water Source Conservation or Protected Areas","O0060020",1,2),
 ("環境用藥管理法施行細則","Enforcement Rules of the Environmental Pesticide Control Act","O0060006",1,2),
 ("環境用藥廣告管理辦法","Regulations Governing Advertising of Environmental Pesticides","O0060031",1,2),
 ("土壤污染評估調查人員管理辦法","Regulations Governing Personnel for Soil Pollution Assessment and Investigation","O0110014",1,2),
 ("土壤污染評估調查及檢測作業管理辦法","Regulations Governing Soil Pollution Assessment, Investigation, and Testing Operations","O0110017",1,2),
 ("污染土地關係人之善良管理人注意義務認定準則","Guidelines for Determining the Duty of Care of a Prudent Administrator for Persons Related to Contaminated Land","O0110021",1,2),
 ("機動車輛排放空氣污染物及噪音檢驗測定機構管理辦法","Regulations Governing Testing Institutions for Motor Vehicle Air Pollutant Emissions and Noise","O0070004",1,2),
 ("環境教育設施場所認證及管理辦法","Regulations Governing the Accreditation and Management of Environmental Education Facilities and Venues","O0120004",1,2),
 ("環境教育機構認證及管理辦法","Regulations Governing the Accreditation and Management of Environmental Education Institutions","O0120005",1,2),
 ("環境檢驗測定機構違反環保法規罰鍰額度裁罰準則","Guidelines for Determining Fine Amounts for Environmental Testing Institutions in Violation of Environmental Protection Regulations","O0070007",1,2),
 ("既存固定污染源污染物排放量認可準則","Approval Guidelines for Pollutant Emission Amounts of Existing Stationary Pollution Sources","O0020064",1,2),
 ("三級防制區既存固定污染源應削減污染物排放量準則","Guidelines for Pollutant Emission Reduction by Existing Stationary Pollution Sources in Class III Control Areas","O0020126",1,2),
 ("機動車輛噪音驗證核可準則","Guidelines for Verification and Approval of Motor Vehicle Noise","O0030015",1,2),
 ("飲用水水質處理藥劑申請公告作業準則","Operational Guidelines for Application and Announcement of Drinking Water Treatment Chemicals","O0040050",1,2),
 ("飲用水水源水質或淨水處理改善計畫審核準則","Guidelines for Review of Drinking Water Source Quality or Water Treatment Improvement Plans","O0040051",1,2),
 ("氣候變遷風險評估作業準則","Guidelines for Climate Change Risk Assessment","O0020143",1,2),
 ("環境用藥原體轉讓申請作業準則","Operational Guidelines for Application for Transfer of Environmental Agent Technical Materials","O0060023",1,2),
 ("環境用藥專供輸出申請作業準則","Operational Guidelines for Application for Environmental Agents Exclusively for Export","O0060026",1,2),
 ("環境用藥專供試驗研究教育示範專案防治或登記用申請作業準則","Operational Guidelines for Application for Environmental Agents Exclusively for Testing, Research, Education, Demonstration Projects, Pest Control, or Registration","O0060027",1,2),
 ("環境用藥許可證申請核發作業準則","Operational Guidelines for Application and Issuance of Environmental Agent Permits","O0060028",1,2),
 ("環境用藥分裝調配及委託製造作業準則","Operational Guidelines for Repackaging, Formulation, and Contract Manufacturing of Environmental Agents","O0060029",1,2),
 ("環境用藥標示準則","Labeling Guidelines for Environmental Agents","O0060030",1,2),
 ("環境用藥販賣業及病媒防治業許可執照申請核發作業準則","Operational Guidelines for Application and Issuance of Permits for Environmental Agent Retailers and Pest Control Operators","O0060032",1,2),
 ("毒性及關注化學物質事故調查處理報告作業準則","Operational Guidelines for Investigation and Reporting of Toxic and Concerned Chemical Substance Incidents","O0060034",1,2),
 ("土壤底泥及地下水污染物檢驗測定品質管制準則","Quality Control Guidelines for Testing of Soil, Sediment, and Groundwater Pollutants","O0110016",1,2),
 ("辦理土壤及地下水污染場址整治目標公聽會作業準則","Operational Guidelines for Public Hearings on Remediation Goals for Soil and Groundwater Pollution Sites","O0110019",1,2),
]

for _i,(_zh,_en,_code,_w) in enumerate(OSH_LAWS, start=1):
    LAWS.setdefault(f"OSH-{_i:02d}", (_zh,_en,"",_code))
for _i,(_zh,_en,_code,_w,_t) in enumerate(ENV_LAWS, start=1):
    LAWS.setdefault(f"ENV-{_i:02d}", (_zh,_en,"",_code))
LAW_NAME = {k: v[0] for k, v in LAWS.items()}
LAW_VER = {k: (v[2] or law_version_from_text(v[0])) for k, v in LAWS.items()}
LAW_EN = {k: v[1] for k, v in LAWS.items()}
LAW_CODE = {k: v[3] for k, v in LAWS.items()}

PCODES = {}
try:
    PCODES = json.load(open(os.path.join(os.path.dirname(HERE), "法規原文", "pcodes.json"), encoding="utf-8"))   # 名稱→全國法規資料庫 pcode（find_pcode.py 查得）
except Exception:
    pass

def law_url(code):
    if not code: return "（尚無法規資料庫代碼）"
    if code.startswith("mol:"): return "https://laws.mol.gov.tw/FLAW/FLAWDAT01.aspx?id=" + code[4:]
    if code.startswith("moi:"): return "https://glrs.moi.gov.tw/LawContent.aspx?id=" + code[4:]
    if code.startswith("isha:"): return "http://law.isha.org.tw/ISHA_LAW/Pages/LawList.aspx?Lawid=" + code[5:]
    return f"https://law.moj.gov.tw/LawClass/LawAll.aspx?pcode={code}"

DELETED_RE = re.compile('^第 ([0-9-]+) 條[^\n]*\n\\s*（刪除）', re.M)   # 標示「（刪除）」的條號
_ART_RE = re.compile(r'^第 [\d\-]+ 條', re.M)
def article_count(zh, lid=None):
    p = os.path.join(os.path.dirname(HERE), "法規原文", zh + ".txt")
    try:
        txt = open(p, encoding="utf-8").read()
        dele = set(DELETED_RE.findall(txt))
        return len([a for a in _ART_RE.findall(txt) if a.replace("第 ", "").replace(" 條", "") not in dele])   # 不計已刪除條，與 coverage.py 一致
    except Exception: return ""

def build_laws(qcount=None):
    qcount = qcount or {}
    rows = [["law_id","group","tier","name_zh","name_en","law_version","source_url","weight","note","articles","questions","family"]]
    byid = {}
    for i,(zh,en,pcode,w) in enumerate(OSH_LAWS, start=1):
        lid = f"OSH-{i:02d}"; code = PCODES.get(zh) or ""
        byid[lid] = [lid,"OSH",1,zh,en,LAW_VER.get(lid) or law_version_from_text(zh),law_url(code),w,"",article_count(zh),qcount.get(lid,0)]
    for i,(zh,en,pcode,w,tier) in enumerate(ENV_LAWS, start=1):
        lid = f"ENV-{i:02d}"; code = PCODES.get(zh) or ""
        byid[lid] = [lid,"ENV",tier,zh,en,law_version_from_text(zh),law_url(code),w,"",article_count(zh),qcount.get(lid,0)]
    order = family_order()
    listed = {lid for lid,_ in order}
    missing = [lid for lid in byid if in_scope(lid) and lid not in listed]
    assert not missing, f"FAMILIES 漏列：{missing}"
    for lid,fam in order:
        if lid in byid and in_scope(lid):
            rows.append(byid[lid] + [fam])
    return rows

# ---------- Questions ----------
Q_HEADER = ["id","law_group","law_id","law","article","law_version","category","difficulty",
            "q_zh","a_zh","b_zh","c_zh","d_zh","q_en","a_en","b_en","c_en","d_en",
            "answer","explain_zh","explain_en","status","batch","reviewer","review_note"]

_DELETED = {}
def deleted_articles(lid):
    """該法規原文中標示「（刪除）」的條號集合（如 {'21','61'}）"""
    if lid in _DELETED: return _DELETED[lid]
    zh = LAWS[lid][0]; p = os.path.join(HERE, "..", "法規原文", zh + ".txt"); out = set()
    if os.path.exists(p):
        txt = open(p, encoding="utf-8").read()
        for m in re.finditer(r'^第 ([\d\-]+) 條[^\n]*\n\s*（刪除）', txt, re.M): out.add(m.group(1))
    _DELETED[lid] = out; return out
def article_keys(art):
    return [n + ('-' + (s1 or s2) if (s1 or s2) else '') for n, s1, s2 in re.findall(r'第\s*(\d+)(?:-(\d+))?\s*條(?:之\s*(\d+))?', str(art))]

DROPPED = {}
DROPPED_DELETED = 0
MIXED_DELETED = []
def load_questions():
    rows = [Q_HEADER]; errs = []; seen = set(); n = 0
    per_batch = {}
    for bno, mods in sorted(BATCHES.items()):
        for m in mods:
            for t in importlib.import_module(m).Q:
                n += 1
                if len(t) != 11: errs.append(f"#{n} 欄位數={len(t)}"); continue
                lid,art,diff,cat,qz,oz,ans,ez,qe,oe,ee = t
                if not in_scope(lid):
                    n -= 1; DROPPED[lid] = DROPPED.get(lid, 0) + 1; continue
                if lid not in LAWS: errs.append(f"#{n} law_id 未知：{lid}"); continue
                ks = article_keys(art)
                if ks and all(k in deleted_articles(lid) for k in ks):      # 已刪除條文的題目不納入（無實質內容）
                    n -= 1; globals()["DROPPED_DELETED"] += 1; continue
                if ks and any(k in deleted_articles(lid) for k in ks):      # 部分條文已刪除：留著但列出來人工看
                    MIXED_DELETED.append(f"{lid} {art}")
                if len(oz)!=4 or len(oe)!=4: errs.append(f"#{n} 選項數不是4：{qz[:20]}")
                if ans not in "abcd": errs.append(f"#{n} 答案非a-d：{ans}")
                if diff not in (1,2,3): errs.append(f"#{n} 難度非1-3")
                if lid not in LAWS: errs.append(f"#{n} law_id 未知：{lid}")
                if len(set(oz))<4 or len(set(oe))<4: errs.append(f"#{n} 選項重複：{qz[:20]}")
                key = (lid, qz.strip())   # 同題幹可跨法規出現（如固定式/移動式起重機同條文）
                if key in seen: errs.append(f"#{n} 題目重複：{qz[:30]}")
                seen.add(key)
                qid = f"Q{n:04d}"
                order = [0,1,2,3]; random.Random(f"{lid}|{qz.strip()}-osh-quiz").shuffle(order)   # 以內容為種子：改一題不會連動改到其他題的答案
                oz2 = [oz[i] for i in order]; oe2 = [oe[i] for i in order]
                ans2 = "abcd"[order.index("abcd".index(ans))]
                row = [qid,lid[:3],lid,LAW_NAME[lid],art,LAW_VER[lid],cat,diff,qz,*oz2,qe,*oe2,ans2,ez,ee,("active" if in_scope(lid) else "archived"),bno,"",("" if in_scope(lid) else "不在使用者指定之職安法規範圍，暫不使用")]
                rows.append(row); per_batch.setdefault(bno, []).append(row)
    return rows, errs, per_batch

CONFIG = [
 ["key","value","說明 / Description"],
 ["questions_per_game",20,"每局題數 / questions per game"],
 ["seconds_per_question",20,"每題秒數 / seconds per question"],
 ["base_score",200,"答對基本分 / base score for a correct answer"],
 ["speed_bonus_max",300,"速度加分上限：base + max×(剩餘秒/總秒) / speed bonus = max × (remaining/total)"],
 ["wrong_score",0,"答錯或逾時得分 / score for wrong or timeout"],
 ["per_question_max",500,"單題最高分（20 題滿分 10000）/ max score per question"],
 ["streak_mult",1.2,"連對時速度分乘數（單題仍不超過 per_question_max）/ streak speed multiplier"],
 ["streak_start",3,"連對從第幾題起加成 / streak bonus starts at N consecutive correct"],
 ["streak_bonus",50,"連對每題加成 / bonus per question once streak active"],
 ["daily_questions",10,"每日挑戰題數 / daily challenge question count"],
 ["lobby_wait_seconds",10,"隨機配對等待秒數，逾時轉房間碼或 bot / matchmaking wait before fallback"],
 ["languages","zh,en","支援語言 / supported languages"],
 ["active_status","active","前端只抓此 status 的題目 / only questions with this status are exported"],
 ["site_mode","active","網站出題模式：draft＝含待審題全部出題；active＝只出 status=active 的題（審完後改這裡即可，不必改程式）/ site question mode"],
]
CHANGELOG = [
 ["date","law_id","law_version","change","affected_questions","action","done_by"],
 ["2026-09-03","OSH-01","民國114年12月19日","建立批次1：以全國法規資料庫現行條文出題（本法）","Q0001–Q0096","初版 draft，待審","Claude Code"],
 ["2026-09-03","OSH-02","民國115年6月26日","建立批次1：以全國法規資料庫現行條文出題（施行細則，115/7/1 施行）","Q0097–Q0129","初版 draft，待審","Claude Code"],
 ["2026-09-03","OSH-07","民國115年6月30日","建立批次2：職業安全衛生設施規則（115/7/1 施行，部分條文 116/1/1）","batch=2 之 OSH-07","初版 draft，待審","Claude Code"],
 ["2026-09-03","OSH-36","民國115年6月30日","建立批次2：營造安全衛生設施標準（第11條之2 自 116/7/1 施行）；Laws 分頁新增 OSH-36","batch=2 之 OSH-36","初版 draft，待審","Claude Code"],
 ["2026-09-03","OSH-13","民國111年5月11日","建立批次2：機械設備器具安全標準；Laws 分頁 OSH-13 來源網址修正為 N0060034","batch=2 之 OSH-13","初版 draft，待審","Claude Code"],
 ["2026-09-03","ENV-26..97","見各題 law_version","建立批次8：環保子法（空污專責人員、固定污染源排放標準、空品標準、室內空品、噪音細則、環境音量標準、水污細則、廢水專責人員、廢清細則、有害事業廢棄物認定、清除處理機構、環評細則、毒化物子法、土污標準、氣候法子法、環教子法…）","batch=8","初版 draft，待審","Claude Code"],
 ["2026-09-03","ENV-29..97, OSH-05/06","見各題 law_version","建立批次9：環保子法逐條出題（空污許可證/空污費/VOC/鍋爐/噪音/水體/飲用水/清除處理/回收再利用/環評認定/毒化物/環境用藥/土水監測/減量交易/碳足跡/環教人員/環檢機構等）＋職災勞工保護法","batch=9","初版 draft，待審","Claude Code"],
 ["2026-09-03","ENV-01..25","見各題 law_version","建立批次10：環保母法逐條出題（空污法/噪音法/水污法/廢清法/環評法/海污法/海岸法/海洋保育法/野保法/濕地法/土污法/毒化法/環境基本法/氣候法/環教法/資源循環推動法/公害糾紛法等）","batch=10","初版 draft，待審","Claude Code"],
 ["2026-09-03","OSH-01..90","見各題 law_version","建立批次11：職安法規逐條補題（每一條文至少一題）","batch=11","初版 draft，待審","Claude Code"],
 ["2026-09-04","ENV-98..207","見各題 law_version","建立批次12：新增 110 部環保細則／標準／準則／管理辦法（空氣、噪音、水、廢棄物、土水、毒化物、氣候、環教）逐條出題","batch=12","初版 draft，待審","Claude Code"],
 ["2026-09-04","ENV-26..96","見各題 law_version","建立批次13：環保子法（批次8 原僅 5–8 題之 28 部）逐條補齊","batch=13","初版 draft，待審","Claude Code"],
 ["2026-09-04","ENV-208..223","見各題 law_version","建立批次14：環保 16 部準則類法規逐條出題","batch=14","初版 draft，待審","Claude Code"],
 ["2026-09-03","OSH-51..96, ENV-26..97","見各題 law_version","建立批次7：依全國法規資料庫「職業安全衛生目／勞動檢查目／環境部各目」總表補齊：新增 46 部職安法規、72 部環保子法（原文已抓齊）；批次7 起建立題目（霸凌準則、工程安全設計、容許暴露、環測、危險性機械設備檢查、化學品三辦法、機械產品申報登錄／型式驗證、健康服務機構、職業病鑑定、職災補助…）","batch=7","初版 draft，待審","Claude Code"],
 ["2026-09-03","OSH-39,40,42,43,44,45,46,47,48,49,50","見各題 law_version","建立批次6：鉛、四烷基鉛、有機溶劑、特定化學物質、粉塵、化學品評估分級、鍋爐壓力容器、起重升降機具、碼頭裝卸、勞動基準法及施行細則（使用者 2026-09-03 增列之職安範圍）","batch=6","初版 draft，待審","Claude Code"],
 ["2026-09-03","OSH-03,04,35,22,20,21,16,17,23,25,29,31,30,19,27,26,33,28","見各題 law_version","建立批次5：勞動檢查法系（含施行細則、立即危險認定標準、違反案件處理要點、重大災害通報要點、優良單位選拔要點）、勞工職業災害保險及保護法、母性健康保護、妊娠及未滿十八歲禁止工作認定標準、女性夜間、精密、重體力、工業用機器人、顧問機構、文化獎勵、績效評核、健檢醫療機構認可、檢驗機構要點；Laws 分頁全部 61 部法規版本日期改由條文檔自動帶入","batch=5","初版 draft，待審","Claude Code"],
 ["2026-09-03","OSH-09,11,10,08,32","見各題 law_version","建立批次4：職業安全衛生管理辦法（115/6/29）、教育訓練規則（115/6/25）、勞工健康保護規則（115/6/26）、危險性工作場所審查及檢查辦法（109/7/17）、製程安全評估定期實施辦法（109/7/17）","batch=4","初版 draft，待審","Claude Code"],
 ["2026-09-03","OSH-34,15,14,18,12,24","見各題 law_version","建立批次3：缺氧症預防規則、高架作業、高溫作業、異常氣壓、危害性化學品標示及通識規則（115/8/26；附表一、四自 118/1/1 施行）、標示設置準則；Laws 分頁 source_url 全面改為法規資料庫查得之正確代碼","batch=3","初版 draft，待審","Claude Code"],
]

def style_header(ws):
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="16324F")
        c.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
def autowidth(ws, maxw=60):
    for col in ws.columns:
        w = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(8, w*1.1), maxw)
def tsv(rows):
    out = []
    for r in rows:
        cells = []
        for c in r:
            s = '' if c is None else str(c)
            cells.append(s.replace('\t',' ').replace('\r',' ').replace('\n',' '))
        out.append('\t'.join(cells))
    return '\n'.join(out)

def _art_key(a):
    """條號排序鍵：第12條之1 → (12,1)；章節名（無條號）→ (10**6, 原字串)"""
    m = re.search(r'第\s*(\d+)(?:-(\d+))?\s*條(?:之\s*(\d+))?', str(a))
    if not m: return (10**6, 0, str(a))
    return (int(m.group(1)), int(m.group(2) or m.group(3) or 0), "")

def sort_and_renumber(qrows, per_batch):
    """職安／環保分開：先職安後環保，各依 Laws 體系順序→法規→條號排序，題號重編為 OSH-0001…／ENV-0001…。
    per_batch 內的列與 qrows 是同一批 list 物件，改 id 會同步反映。"""
    order = {lid: i for i, (lid, _f) in enumerate(family_order())}
    head, rows = qrows[0], qrows[1:]
    rows.sort(key=lambda r: (0 if r[1] == "OSH" else 1, order.get(r[2], 10**6), _art_key(r[4])))
    n = {"OSH": 0, "ENV": 0}
    for r in rows:
        n[r[1]] += 1; r[0] = f"{r[1]}-{n[r[1]]:04d}"
    return [head] + rows

def main():
    qrows, errs, per_batch = load_questions()
    if errs:
        print("驗證錯誤："); [print(" ", e) for e in errs]; sys.exit(1)
    qrows = sort_and_renumber(qrows, per_batch)
    from collections import Counter
    laws_rows = build_laws(Counter(r[2] for r in qrows[1:]))
    wb = Workbook()
    split = {}
    for g in ("OSH", "ENV"):
        split[g] = ([laws_rows[0]] + [r for r in laws_rows[1:] if r[1] == g],
                    [qrows[0]] + [r for r in qrows[1:] if r[1] == g])
    ws = wb.active; ws.title = "Laws_OSH"
    for r in split["OSH"][0]: ws.append(r)
    style_header(ws); autowidth(ws)
    ws = wb.create_sheet("Laws_ENV")
    for r in split["ENV"][0]: ws.append(r)
    style_header(ws); autowidth(ws)
    ws = wb.create_sheet("Questions_OSH")
    for r in split["OSH"][1]: ws.append(r)
    style_header(ws); autowidth(ws, 50)
    ws = wb.create_sheet("Questions_ENV")
    for r in split["ENV"][1]: ws.append(r)
    style_header(ws); autowidth(ws, 50)
    ws = wb.create_sheet("Config")
    for r in CONFIG: ws.append(r)
    style_header(ws); autowidth(ws, 80)
    ws = wb.create_sheet("Changelog")
    for r in CHANGELOG: ws.append(r)
    style_header(ws); autowidth(ws, 70)
    wb.save(os.path.join(HERE, "OSH_ENV_QuizBank.xlsx"))
    keys = qrows[0]
    js = [dict(zip(keys, r)) for r in qrows[1:]]
    json.dump(js, open(os.path.join(HERE,"questions_all.json"),"w",encoding="utf-8"), ensure_ascii=False, indent=1)
    os.makedirs(os.path.join(HERE,"tsv"), exist_ok=True)
    for bno, rows in per_batch.items():
        json.dump([dict(zip(keys, r)) for r in rows], open(os.path.join(HERE,f"questions_batch{bno}.json"),"w",encoding="utf-8"), ensure_ascii=False, indent=1)
        open(os.path.join(HERE,"tsv",f"Questions_batch{bno}.tsv"),"w",encoding="utf-8",newline="").write(tsv(rows))
    open(os.path.join(HERE,"tsv","Questions.tsv"),"w",encoding="utf-8",newline="").write(tsv(qrows))   # 全部題目（含表頭），push_to_sheet.py sync 用
    open(os.path.join(HERE,"tsv","Laws.tsv"),"w",encoding="utf-8",newline="").write(tsv(laws_rows))
    for g in ("OSH", "ENV"):
        open(os.path.join(HERE,"tsv",f"Laws_{g}.tsv"),"w",encoding="utf-8",newline="").write(tsv(split[g][0]))
        open(os.path.join(HERE,"tsv",f"Questions_{g}.tsv"),"w",encoding="utf-8",newline="").write(tsv(split[g][1]))
    open(os.path.join(HERE,"tsv","Changelog.tsv"),"w",encoding="utf-8",newline="").write(tsv(CHANGELOG))
    open(os.path.join(HERE,"tsv","Config.tsv"),"w",encoding="utf-8",newline="").write(tsv(CONFIG))
    # 前端用精簡 JSON
    keep = ['id','law_id','law','article','law_version','category','difficulty','q_zh','a_zh','b_zh','c_zh','d_zh','q_en','a_en','b_en','c_en','d_en','answer','explain_zh','explain_en','status']
    js = [q for q in js if q.get('status') != 'archived']
    docs = os.path.join(os.path.dirname(HERE), "docs", "data", "questions.json")
    if os.path.isdir(os.path.dirname(docs)):
        json.dump({'generated': __import__('datetime').date.today().isoformat(),'count':len(js),'questions':[{k:r[k] for k in keep} for r in js]}, open(docs,"w",encoding="utf-8"), ensure_ascii=False, separators=(',',':'))
    n = len(qrows)-1
    print(f"已略過已刪除條文題目：{DROPPED_DELETED} 題")
    if DROPPED: print("不在範圍略過：", dict(DROPPED), "共", sum(DROPPED.values()), "題")     # P2
    if MIXED_DELETED: print("引用條文含已刪除條（請人工確認）：", len(MIXED_DELETED), MIXED_DELETED[:8])
    print(f"OK：{n} 題（{', '.join(f'批次{b}={len(r)}' for b,r in sorted(per_batch.items()))}）")
    print("依法規：", dict(Counter(r[2] for r in qrows[1:])))
    print("依難度：", dict(sorted(Counter(r[7] for r in qrows[1:]).items())))
    print("答案分布：", dict(sorted(Counter(r[18] for r in qrows[1:]).items())))
    for g in ("OSH", "ENV"):
        rs = split[g][1][1:]
        print(f"{g} 題號 {rs[0][0]}–{rs[-1][0]}，共 {len(rs)} 題，答案分布", dict(sorted(Counter(r[18] for r in rs).items())))

if __name__ == "__main__":
    main()
